import random
import openpyxl
from datetime import datetime, timedelta

# 定义生成数据的参数
num_days = 7     # 生成 7 天的数据
records_per_day = 4  # 每天生成 4 条记录
min_wind_speed = 2   # 最低风速为 2 m/s
max_wind_speed = 20  # 最高风速为 20 m/s
wind_directions = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Wind Data"

# 写入表头
worksheet["A1"] = "Date"
worksheet["B1"] = "Time"
worksheet["C1"] = "Wind Direction"
worksheet["D1"] = "Wind Speed (m/s)"

# 生成并写入随机风向和风速数据
start_date = datetime(2024, 5, 20)
for day in range(num_days):
    for record in range(records_per_day):
        row = day * records_per_day + record + 2
        date_cell = worksheet.cell(row=row, column=1)
        date_cell.value = start_date + timedelta(days=day)
        
        time_cell = worksheet.cell(row=row, column=2)
        time_cell.value = f"{record*6:02d}:00"
        
        wind_direction_cell = worksheet.cell(row=row, column=3)
        wind_direction_cell.value = random.choice(wind_directions)
        
        wind_speed_cell = worksheet.cell(row=row, column=4)
        wind_speed_value = round(random.uniform(min_wind_speed, max_wind_speed), 1)
        wind_speed_cell.value = wind_speed_value

# 保存 Excel 文件
workbook.save("wind_data.xlsx")
print("风向风速数据已成功保存到 wind_data.xlsx 文件中。")