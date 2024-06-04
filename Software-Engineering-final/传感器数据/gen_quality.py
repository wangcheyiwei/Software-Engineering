import random
import openpyxl
from datetime import datetime, timedelta

# 定义生成数据的参数
num_days = 7     # 生成 7 天的数据
records_per_day = 4  # 每天生成 4 条记录
min_aqi = 20     # 最低 AQI 为 20
max_aqi = 150    # 最高 AQI 为 150
aqi_categories = ["优", "良", "中", "较差", "极差"]

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Air Quality Data"

# 写入表头
worksheet["A1"] = "Date"
worksheet["B1"] = "Time"
worksheet["C1"] = "AQI"
worksheet["D1"] = "AQI Category"

# 生成并写入随机空气质量数据
start_date = datetime(2024, 5, 20)
for day in range(num_days):
    for record in range(records_per_day):
        row = day * records_per_day + record + 2
        date_cell = worksheet.cell(row=row, column=1)
        date_cell.value = start_date + timedelta(days=day)
        
        time_cell = worksheet.cell(row=row, column=2)
        time_cell.value = f"{record*6:02d}:00"
        
        aqi_cell = worksheet.cell(row=row, column=3)
        aqi_value = random.randint(min_aqi, max_aqi)
        aqi_cell.value = aqi_value
        
        aqi_category_cell = worksheet.cell(row=row, column=4)
        aqi_category = aqi_categories[min(int(aqi_value / 50), 4)]
        aqi_category_cell.value = aqi_category

# 保存 Excel 文件
workbook.save("air_quality_data.xlsx")
print("空气质量数据已成功保存到 air_quality_data.xlsx 文件中。")