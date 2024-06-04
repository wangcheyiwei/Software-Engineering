import random
import openpyxl
from datetime import datetime, timedelta

# 定义生成数据的参数
num_days = 7     # 生成 7 天的数据
records_per_day = 4  # 每天生成 4 条记录
min_temp = 10   # 最低温度为 10°C
max_temp = 30   # 最高温度为 30°C

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Temperature Data"

# 写入表头
worksheet["A1"] = "Date"
worksheet["B1"] = "Time"
worksheet["C1"] = "Temperature (°C)"

# 生成并写入随机温度数据
start_date = datetime(2024, 5, 20)
for day in range(num_days):
    for record in range(records_per_day):
        row = day * records_per_day + record + 2
        date_cell = worksheet.cell(row=row, column=1)
        date_cell.value = start_date + timedelta(days=day)
        
        time_cell = worksheet.cell(row=row, column=2)
        time_cell.value = f"{record*6:02d}:00"
        
        temp_cell = worksheet.cell(row=row, column=3)
        temp_value = round(random.uniform(min_temp, max_temp), 1)
        temp_cell.value = temp_value

# 保存 Excel 文件
workbook.save("temperature_data.xlsx")
print("温度数据已成功保存到 temperature_data.xlsx 文件中。")