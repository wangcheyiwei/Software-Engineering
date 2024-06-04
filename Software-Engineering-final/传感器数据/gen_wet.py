import random
import openpyxl
from datetime import datetime, timedelta

# 定义生成数据的参数
num_days = 7     # 生成 7 天的数据
records_per_day = 4  # 每天生成 4 条记录
min_humidity = 30   # 最低湿度为 30%
max_humidity = 90   # 最高湿度为 90%

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Humidity Data"

# 写入表头
worksheet["A1"] = "Date"
worksheet["B1"] = "Time"
worksheet["C1"] = "Humidity (%)"

# 生成并写入随机湿度数据
start_date = datetime(2024, 5, 20)
for day in range(num_days):
    for record in range(records_per_day):
        row = day * records_per_day + record + 2
        date_cell = worksheet.cell(row=row, column=1)
        date_cell.value = start_date + timedelta(days=day)
        
        time_cell = worksheet.cell(row=row, column=2)
        time_cell.value = f"{record*6:02d}:00"
        
        humidity_cell = worksheet.cell(row=row, column=3)
        humidity_value = round(random.uniform(min_humidity, max_humidity), 1)
        humidity_cell.value = humidity_value

# 保存 Excel 文件
workbook.save("humidity_data.xlsx")
print("湿度数据已成功保存到 humidity_data.xlsx 文件中。")